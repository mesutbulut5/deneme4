import openpyxl

# Excel dosyasının yolu
excel_file = 'C:/Users/mesutbulut/Desktop/MGM/Nem.xlsx'

# Excel dosyasını yükle
workbook = openpyxl.load_workbook(excel_file)

# 'İstasyon' kelimesini aramak için tüm sayfaları kontrol et
search_term = 'İstasyon'

# Tüm sayfaları dolaş
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    print(f"Arama yapılıyor: {sheet_name}")

    # Tüm hücrelerde 'İstasyon' kelimesini ara
    for row in sheet.iter_rows():
        for cell in row:
            if search_term in str(cell.value):  # Eğer hücrede 'İstasyon' kelimesi varsa
                value = str(cell.value)

                # "İstasyon Adı/No:" kısmından sonrasını al
                if "İstasyon Adı/No:" in value:
                    after_colon = value.split("İstasyon Adı/No:")[1].strip()  # ':' sonrasını al
                    # ':' ile '/' arasındaki kısmı al
                    station_name = after_colon.split('/')[0].strip()
                    print(f"Bulunan İstasyon Adı: {station_name}")
